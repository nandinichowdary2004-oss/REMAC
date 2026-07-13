import pandas as pd
import time
import json
import requests
from pathlib import Path
from openpyxl import Workbook, load_workbook

from awscrt import mqtt
from awsiot import mqtt_connection_builder


# =======================
# AWS CONFIG
# =======================

ENDPOINT = "a1kneu9xpfe402-ats.iot.eu-north-1.amazonaws.com"
CLIENT_ID = "REMAC-Node-1"
TOPIC = "remac/node1/data"

ROOT_CA = "certificates/Remac-Node-1/AmazonRootCA1.pem"
CERT = "certificates/Remac-Node-1/dac48685c1e6abf330c38be199ba904aa71eebd042d0d78ec0685dd5e06f8909-certificate.pem.crt"
KEY = "certificates/Remac-Node-1/dac48685c1e6abf330c38be199ba904aa71eebd042d0d78ec0685dd5e06f8909-private.pem.key"


# =======================
# FILE PATHS
# =======================

csv_file = Path("datasets/Node1.csv")
excel_file = Path("excel_logs/Node1_Output.xlsx")


# =======================
# AWS CONNECT
# =======================

print("Connecting to AWS IoT Core...")

mqtt_connection = mqtt_connection_builder.mtls_from_path(
    endpoint=ENDPOINT,
    cert_filepath=CERT,
    pri_key_filepath=KEY,
    ca_filepath=ROOT_CA,
    client_id=CLIENT_ID,
    clean_session=False,
    keep_alive_secs=30,
)

mqtt_connection.connect().result()

print("Connected to AWS IoT Core")


# =======================
# LOAD CSV
# =======================

df = pd.read_csv(csv_file)


# =======================
# SETUP EXCEL
# =======================

if excel_file.exists():
    wb = load_workbook(excel_file)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "Node1 Data"
    ws.append(list(df.columns))  # headers


# =======================
# MAIN LOOP
# =======================

for index, row in df.iterrows():

    data = row.to_dict()

    # 1. Write to Excel
    ws.append(list(data.values()))
    wb.save(excel_file)

    # Convert row data to JSON payload
    payload = json.dumps(data)

    # 2. Publish raw sensor data to AWS IoT Core
    mqtt_connection.publish(
        topic=TOPIC,
        payload=payload,
        qos=mqtt.QoS.AT_LEAST_ONCE
    )

    print(f"Published Row {index + 1} to AWS IoT Core")

    # 3. Wait 5 seconds
    time.sleep(5)


# =======================
# DISCONNECT
# =======================

mqtt_connection.disconnect().result()

print("Simulation Completed")